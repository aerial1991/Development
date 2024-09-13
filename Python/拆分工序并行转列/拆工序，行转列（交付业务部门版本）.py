import tkinter as tk
from tkinter import messagebox

def main():
    # 在这里编写你的程序
    import pandas as pd
    import openpyxl
    from tkinter import filedialog

    # 选择一个文件
    # file_path = filedialog.askopenfilename(title='请选择一个Excel表', filetypes=[('Excel','.xls .xlsx'),('All Files', '*')],initialdir='E:\\')
    # 选择多个文件
    file_path_turple = filedialog.askopenfilename(title='请选择一个Excel表',
                                                  filetypes=[('Excel', '.xls .xlsx'), ('文本', '.txt'),
                                                             ('All Files', '*')], initialdir='E:\\', multiple=True)

    file_path = file_path_turple[0]
    print(file_path)

    # 读取Excel文件
    def row_to_column(path, sheet_name):
        df = pd.read_excel(path, sheet_name=sheet_name)
        # re = {}
        result = []

        # df['物料编码'] = df['物料编码'].astype(str)

        for index, row in df.iterrows():
            # print(row)
            cell_value1 = str(row['物料编码'])  # 第一列的数据
            cell_value3 = row['手工码']  # 第三列的数据
            print(cell_value1)
            # 判断物料编码是否0开头
            if cell_value1.startswith('0'):
                cell_value1 = cell_value1
            else:
                cell_value1 = '0' + cell_value1
            cell_value2 = str(row['工序集'])  # 第二列的数据
            # print(cell_value2)
            list_cell_value2 = cell_value2.split('-')
            # print( list_cell_value2)
            # 增加工序顺序号
            i = 0
            for item in list_cell_value2:
                i += 1
                result.append([cell_value1, item, i * 10, cell_value3])
        print(result)
        return result

    result = row_to_column(file_path, 'Sheet1')
    workbook = openpyxl.load_workbook(file_path)

    # 获取所有sheet页的名称
    sheet_names = workbook.sheetnames
    print(sheet_names)
    # 判断'Sheet3' 是否存在
    if 'Result' in sheet_names:
        sheet = workbook['Result']
    else:
        workbook.create_sheet('Result')
        sheet = workbook['Result']

    # 写入数据
    sheet['A1'] = '物料编码'
    sheet['B1'] = '工序'
    sheet['C1'] = '顺序号'
    sheet['D1'] = '手工码'

    for row in result:
        sheet.append(row)
    workbook.save(file_path)


if __name__ == "__main__":
    main()
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("提示", "操作成功！")
