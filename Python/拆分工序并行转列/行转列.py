import pandas as pd
import openpyxl
# 读取Excel文件
def row_to_column(path,sheet_name):
    df = pd.read_excel(path,sheet_name=sheet_name)
    # re = {}
    result = []

    df['物料编码'] = df['物料编码'].astype(str)

    for index, row in df.iterrows():
        # print(row)
        cell_value1 = row['物料编码'] # 第一列的数据
        #判断物料编码是否0开头
        if cell_value1.startswith('0'):
            cell_value1=cell_value1
        else:
            cell_value1 ='0'+cell_value1
        cell_value2 = row['工序集'] # 第二列的数据
        # print(cell_value2)
        list_cell_value2 =cell_value2.split('-')
        # print( list_cell_value2)
        for item in list_cell_value2:
            result.append([cell_value1,item])
    print(result)
    return result

result =row_to_column('D:/111.xlsx','Sheet2')
workbook = openpyxl.load_workbook('D:/111.xlsx')


# 获取所有sheet页的名称
sheet_names = workbook.sheetnames
print(sheet_names)
#判断'Sheet3' 是否存在
if 'Sheet3' in sheet_names:
    sheet = workbook['Sheet3']
else:
    workbook.create_sheet('Sheet3')
    sheet = workbook['Sheet3']

# 写入数据
sheet['A1'] = '物料编码'
sheet['B1'] = '工序'
for row in result:
    sheet.append(row)
workbook.save('D:/111.xlsx')
