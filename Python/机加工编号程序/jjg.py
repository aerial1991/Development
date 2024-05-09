import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QAction, QMessageBox,QLabel
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSlot
from PyQt5.Qt import QLineEdit
from PyQt5.Qt import QComboBox
from PyQt5.Qt import QFont
import openpyxl
from datetime import datetime


class App(QWidget):

    def __init__(self):
        super().__init__()
        self.title = '样品编码程序'
        self.left = 500
        self.top = 100
        self.width = 800
        self.height = 1000
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        # create label
        self.label = QLabel(self)
        self.label.setText("原始编码:")
        self.label.setGeometry(20, 40, 200, 30)
        self.label.setFont(QFont("Roman times",20,QFont.Bold))

        # create labe2
        self.label2 = QLabel(self)
        self.label2.setText("项目编码:")
        self.label2.setGeometry(20, 120, 200, 30)
        self.label2.setFont(QFont("Roman times", 20, QFont.Bold))
        # create labe3
        self.label3 = QLabel(self)
        self.label3.setText("数   量 :")
        self.label3.setGeometry(20, 200, 200, 30)
        self.label3.setFont(QFont("Roman times", 20, QFont.Bold))
        # create labe4
        self.label4 = QLabel(self)
        self.label4.setText("提示：SL室温拉伸试件 GL高温拉伸试件 ZL Z向拉伸试件 CJ冲击试件 WQ弯曲试件")
        self.label4.setGeometry(20, 410, 2000, 200)

        # create labe5
        self.label5 = QLabel(self)
        self.label5.setText("样品编号:")
        self.label5.setGeometry(20, 370, 200, 30)
        self.label5.setFont(QFont("Roman times", 20, QFont.Bold))
        # create labe6
        self.label6 = QLabel(self)
        self.label6.setText("  KK扩口试件 HX化学试件 WJ微观金相试件 HJ宏观金相试件 YD硬度试件 QT其他试件")
        self.label6.setGeometry(20, 430, 2000, 200)

        # create textbox1
        self.textbox = QLineEdit(self)
        self.textbox.move(200, 20)
        self.textbox.resize(400, 70)
        self.textbox.setFont(QFont("Roman times", 20, QFont.Bold))

        # create textbox2
        # self.textbox2 = QLineEdit(self)
        # self.textbox2.move(100, 100)
        # self.textbox2.resize(280, 40)

        items = ["CJ", "SL", "HX","GL", "KK", "ZL", "WQ", "YB",   "WJ", "HJ", "YD", "QT"]
        self.comboBox = QComboBox(self)
        self.comboBox.addItems(items)
        self.comboBox.setCurrentIndex(1)  # 设置默认值
        self.comboBox.currentText()  # 获得当前内容
        self.comboBox.move(200,100)
        self.comboBox.resize(400, 70)
        self.comboBox.setFont(QFont("Roman times", 20, QFont.Bold))
        # create textbox3
        self.textbox3 = QLineEdit(self)
        self.textbox3.move(200, 180)
        self.textbox3.resize(400, 70)
        self.textbox3.setFont(QFont("Roman times", 20, QFont.Bold))
        # create textbox4
        self.textbox4 = QLineEdit(self)
        self.textbox4.move(200, 350)
        self.textbox4.resize(400, 70)
        self.textbox4.setFont(QFont("Roman times", 20, QFont.Bold))

        # Create a button in the window
        self.button = QPushButton('确 定', self)
        self.button.move(20, 280)
        self.button.resize(130, 60)
        self.button.setFont(QFont("Roman times", 20, QFont.Bold))
        # self.button.setStyleSheet("background-color: rgb(135, 206, 235);border: 1px solid black;")
        # self.button.setStyleSheet('background-color: rgb(135, 206, 235);border-radius: 10px; ')
        self.button.setStyleSheet("QPushButton:hover {background-color: grey;}QPushButton:pressed {background-color: red;} ")
        self.button.setShortcut("Alt+G")
        self.button.clicked.connect(self.on_click)
        self.show()


    def on_click(self):

        code = self.textbox.text()
        sheet_name = self.comboBox.currentText()
        num_copies = self.textbox3.text()
        if code == "" or num_copies == "":
            code = 0
            num_copies = 0
         # 打开Excel文件
        workbook = openpyxl.load_workbook(r'C:/Users/PT/Desktop/试件编码/试件编码.xlsx')
        
        # 选择指定的sheet页
        sheet = workbook[sheet_name]

        
        count = sheet.max_row
        print(count)
        max_row = sheet.max_row                    

        for i in range(1, int(num_copies)+1):
            new_row = max_row + i
            new_value = sheet_name + str(new_row-1).zfill(5)
            sheet.cell(row=new_row, column=1).value = new_value   
            
                 
        z1 = []
        for i in range(int(num_copies)):
            count = count + 1            
            sheet.cell(row=count, column=2, value=code)
            sheet.cell(row=count, column=3, value=datetime.now())
            z2 = sheet.cell(count, 1).value
            # print(z2)
            z1.append(z2)
        # print(z1)
        self.textbox4.setText(" ".join(z1))

        workbook.save(r'C:/Users/PT/Desktop/试件编码/试件编码.xlsx')

if __name__ == '__main__':
    
    app = QApplication(sys.argv)
    ex = App()
    app.exit(app.exec_())
    


