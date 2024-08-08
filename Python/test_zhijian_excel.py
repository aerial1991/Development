import openpyxl
import pandas as pd

def test1():
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook('/Users/wangyang/Downloads/123.xlsx')
    # 获取第一个工作表
    worksheet = workbook['Sheet1']
    # 创建空列表，用于储存第L列的内容
    column2 = []
    # 遍历工作表的每一行
    for row in worksheet.iter_rows():
        if '工艺增加' in str(row[11].value):
            column2.append(row[0].row)
    #将list倒序排列
    column2.sort(reverse=True)
    #循环遍历list
    for n in column2:
        #倒着删除
        worksheet.delete_rows(n,1)

    # 查看C列，名称包含锅筒或键位器，且E列等于部件，获取A列的值存放到一个集合
    part_set = set()
    for row in worksheet.iter_rows():
        if '集箱' in row[2].value or '简温器' in row[2].value:
            if row[4].value == '部件':
                part_set.add(row[0].value)
                print(str(part_set)+'-------')

    # 创建sheet2页
    sheet2 = workbook.create_sheet('Sheet2')

    # 检索整张表的A列以集合开头的数据，复制到sheet2页中
    row_num = 1
    for row in worksheet.iter_rows():
        if row[0].value.startswith(tuple(part_set)):
            if 't' in str(row[10].value):
                for cell in row:
                    sheet2.cell(row=row_num, column=cell.column, value=cell.value)
                row_num += 1
    
    # 保存Excel表格A
    workbook.save('/Users/wangyang/Downloads/123.xlsx')
    
    test2()

def test2():
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook('/Users/wangyang/Downloads/123.xlsx')

    # 获取第二个工作表
    worksheet2 = workbook['Sheet2']

    # 保存需要复制的数据的字典
    copy_data = {}
    # 遍历 Sheet2 中的每一行
    for row in worksheet2.iter_rows(min_row=1, values_only=True):

        # 判断是否符合条件
        if ('锅筒' in row[2] or '简温器' in row[2]) and row[4] == '部件':
            key = row[0]

            # 如果已经存在对应键值，则将复制次数加 1
            if key in copy_data:
                copy_data[key] += row[3]
            else:
                copy_data[key] = row[3]

    # 在 Sheet2 中以 A 列的值开头的位置插入数据
    for row in workbook.active.iter_rows(min_row=1, values_only=True):
        if row[0] in copy_data:
            key = row[0]
            count = copy_data[key]

            # 复制数据
            for i in range(int(count)):
                new_row = [key] + list(row[1:])
                worksheet2.append(new_row)

    # 保存Excel表格
    workbook.save('/Users/wangyang/Downloads/123.xlsx')

if __name__ == '__main__':
    test1()


