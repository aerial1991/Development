import datetime
import openpyxl
import pymysql
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


# 从数据库某个表中取出所有记录
# 参数host指定数据库服务器的IP地址，参数db_name指定数据库的名字，
# 参数table_name指定表的名字
# 参数user指定数据库的登录用户名，参数passwd指定登录用户的密码
def get_data(host, db_name, table_name, user, passwd):
    # 生成一个数据库的连接
    conn = pymysql.connect(host=host, port=3306, database=db_name, user=user, passwd=passwd)
    # 建立一个游标
    cur = conn.cursor()
    # 组合一个SQL查询语句
    sql = """SELECT
    	pa.CODE AS 物料编码,
    	pa.NAME AS 名称,
    	pa.model AS 型号,
    	pa.figure_no AS 图号,
    	pa.specification AS 规格,
    	pa.material AS 材料,
    	def.define1 AS 执行标准,
    	pa.part_type AS 零部件类型,
    	ba.NAME AS 业务属性,
    	cgr.CODE AS 物料类型编码,
    	tkfile.name_attribute AS 名称属性,
    	pa.creationtime AS 创建时间,
    	ur.name as 创建人
    FROM
    	imp_plm_p_m_part pa
    	LEFT JOIN imp_plm_p_m_def def ON pa.id = def.id
    	LEFT JOIN imp_plm_p_m_cgr cgr ON pa.partclassid = cgr.id
    	LEFT JOIN imp_plm_p_busattr_base ba ON pa.selfmade = ba.id
    	LEFT JOIN tkkgl_db.imp_plm_p_b_l_file tkfile ON pa.name_attribute = tkfile.id
    	LEFT JOIN iuap_uuas_usercenter.USER ur ON pa.creator = ur.yht_user_id 	
    WHERE
    	pa.dr = 0 
    	AND pa.ytenant_id = 'jpgfp4ia' 
    	AND pa.lifecyclestate != '1613019747583525147'
    	and ur.yht_tenant_id = 'jpgfp4ia'"""

    # 执行SQL语句
    cur.execute(sql)
    # rows取得记录，cur.fetchall()返回所有符合条件的记录
    rows = cur.fetchall()
    # cur.description返回数据表的字段信息，
    # 返回值fields是一个元组，其中的每一项元素也是一个元组（子元组），
    # 这个子元组的第一个元素是字段名
    fields = cur.description
    # print(fields)
    # 关闭游标
    cur.close()
    # 断开连接
    conn.close()
    return fields, rows

# 检查有没有异常字符
def check_string( value):
    if type(value) is not str:
        value = str(value)
    """Check string coding, length, and line break character"""

    # convert to str string
    # if not isinstance(value, str):
    #     value = str(value, self.encoding)
    # value = str(value)
    # string must never be longer than 32,767 characters
    # truncate if necessary
    # value = value[:32767]
    if next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
          # 如果发现包含非法字符，则替换为空。
        value = ILLEGAL_CHARACTERS_RE.sub(r'', value)
        # raise IllegalCharacterError**
    if value == "None":
        value = ""


    return value



# 将表的记录导入到Excel中的函数
# 参数host指定数据库服务器的IP地址，参数db_name指定数据库的名字，
# 参数table_name指定表的名字，参数user指定数据库的登录用户名，
# 参数passwd指定登录用户的密码，参数filename指定导入的Excel文件名
def export_to_excel(fields, table_rows, filename):
    # 调用函数，取得数据表的字段信息和记录信息
    # 生成Excel文件的工作簿
    workbook = openpyxl.Workbook()
    # 在工作簿中生成一个工作表，表名设为"table_"加数据表名
    sheet = workbook.create_sheet('table_' + table_name, 1)
    # 在工作表第1行上写上字段名
    for i in range(0, len(fields)):
        # 在openpyxl模块中定义工作表的行始值是1，列起始值是1，
        # 所以cell()函数第1个参数是1表示第1行，第二参数为i+1是因为i从0开始计数,
        # fields[i][0]取得字段的名称
        sheet.cell(1, i + 1, fields[i][0])
    # 从工作表第2行开始写入每条记录的内容
    for row in range(0, len(table_rows)):
        for col in range(0, len(fields)):
            value =check_string(table_rows[row][col])
            sheet.cell(row + 2, col + 1, '%s' % value)
    # 保存到Excel文件中
    workbook.save(filename)


# 主函数main
if __name__ == '__main__':
    # 初始化各变量值
    host = '10.97.0.17'# 数据库服务器的IP地址
    db_name = 'imp_plm' # 数据库的名字
    table_name = 'output111' # 指定数据库中表的名字
    user = 'custom_readonly' # 用户名
    password = 'qQRek8q9tuykzCRz' # 密码
    # 调用函数，将数据表的内容导入到一个Excel文件中
    fields, table_rows = get_data(host, db_name, table_name, user, password)
    # print("fields", fields)
    print("table_rows", table_rows)
    print("table_rows", type(table_rows[0][11]))
    export_to_excel(fields,table_rows,'./研发物料.xlsx')
